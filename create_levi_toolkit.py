import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建 workbook
wb = openpyxl.Workbook()

# 定义样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
subheader_font = Font(bold=True, size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Sheet 1: 交易日志模板
ws1 = wb.active
ws1.title = '交易日志'

headers1 = ['日期', '星期', '品种', '方向', '周期', '开仓时间', '开仓价', '止损价', '止损跳数',
            '仓位(手)', '开仓理由', '平仓时间', '平仓价', '盈亏金额', '盈亏(跳)', '手续费',
            '净利润', '平仓原因', '是否按计划', '情绪状态', '复盘 notes']
ws1.append(headers1)
for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# 添加示例数据
example1 = ['2026-07-15', '二', 'SA2509', '多', '1分钟', '09:35', '1450', '1447', '3',
            '2', '突破开盘平台+放量', '09:42', '1455', '1000', '5', '50', '950', '达到目标位', '是', '平静', '流畅']
ws1.append(example1)

# 设置列宽
column_widths1 = [10, 6, 10, 6, 8, 10, 10, 10, 10, 10, 25, 10, 10, 12, 10, 10, 12, 15, 10, 10, 25]
for i, width in enumerate(column_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# 添加汇总区域
ws1['A25'] = '月度汇总'
ws1['A25'].font = Font(bold=True, size=12)
ws1['A26'] = '总交易次数'
ws1['B26'] = '=COUNTA(A2:A24)'
ws1['A27'] = '盈利次数'
ws1['B27'] = '=COUNTIF(N2:N24,">0")'
ws1['A28'] = '亏损次数'
ws1['B28'] = '=COUNTIF(N2:N24,"<0")'
ws1['A29'] = '胜率'
ws1['B29'] = '=B27/B26'
ws1['B29'].number_format = '0.00%'
ws1['A30'] = '总盈亏'
ws1['B30'] = '=SUM(N2:N24)'
ws1['A31'] = '总净利润(扣手续费)'
ws1['B31'] = '=SUM(Q2:Q24)'
ws1['A32'] = '平均单笔盈亏'
ws1['B32'] = '=AVERAGE(N2:N24)'
ws1['A33'] = '最大单笔盈利'
ws1['B33'] = '=MAX(N2:N24)'
ws1['A34'] = '最大单笔亏损'
ws1['B34'] = '=MIN(N2:N24)'

for row in range(26, 35):
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'A{row}'].fill = subheader_fill

# Sheet 2: 假突破识别清单
ws2 = wb.create_sheet('假突破识别清单')

ws2['A1'] = '假突破识别检查清单'
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:E1')
ws2['A1'].alignment = center_align

headers2 = ['序号', '假突破信号', '具体表现', '危险等级', '应对方式']
ws2.append(headers2)
for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

fake_breakouts = [
    ['1', '无量突破', '突破时成交量没有明显放大，K线实体小', '高', '观望，等回踩确认再进'],
    ['2', '尾盘/收盘前突击突破', '临近收盘突然拉一波，缺乏持续性', '高', '不参与，等第二天确认'],
    ['3', '长上影线/长下影线', '突破后迅速被拉回，留下长影线', '高', '放弃这次突破'],
    ['4', '毛刺式突破', '价格快速刺破关键位又立刻回来', '高', '等实体站稳再进'],
    ['5', '已经涨/跌了一大段后的突破', '趋势末端加速，空间不大', '中', '减仓参与或放弃'],
    ['6', '消息面刺激的假突破', '突发消息导致跳空，随后快速回落', '高', '等消息消化再判断'],
    ['7', '多次测试同一位置后突破', '同一位置反复测试，筹码松动', '中', '等突破后回踩不破再进'],
    ['8', '逆势突破', '与当天主趋势相反方向的突破', '高', '只顺应当日主方向'],
    ['9', '小级别突破但大级别没配合', '1分钟突破，但5分钟/15分钟没方向', '中', '降低仓位或放弃'],
    ['10', '突破后没有加速', '突破后横盘、犹豫、缺乏跟进', '中', '减仓或离场'],
    ['11', '整数关口假突破', '价格突破整数位吸引跟风盘后回落', '中', '等突破整数位一定距离再进'],
    ['12', '开盘跳空突破', '开盘直接跳空突破，随后回补缺口', '中', '等缺口不回补再追'],
]

for row in fake_breakouts:
    ws2.append(row)

# 设置样式
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=5):
    for cell in row:
        cell.border = thin_border
        cell.alignment = left_align
        if cell.column == 4:  # 危险等级
            if cell.value == '高':
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                cell.font = Font(color='9C0006')
            elif cell.value == '中':
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                cell.font = Font(color='9C5700')

column_widths2 = [6, 18, 35, 10, 30]
for i, width in enumerate(column_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# Sheet 3: 仓位管理对照表
ws3 = wb.create_sheet('仓位管理对照表')

ws3['A1'] = 'Levi 风格仓位管理对照表'
ws3['A1'].font = Font(bold=True, size=14)
ws3.merge_cells('A1:F1')
ws3['A1'].alignment = center_align

ws3['A3'] = '基本原则'
ws3['A3'].font = Font(bold=True, size=12)
ws3['A4'] = '单笔亏损不超过总资金的 1%'
ws3['A5'] = '单日亏损不超过总资金的 3%'
ws3['A6'] = '同一个品种最多给 2 次机会'
ws3['A7'] = '连续亏损 3 次后暂停交易'
ws3['A8'] = '盈利后循序渐进加仓，不要跳跃式放大'

headers3 = ['资金规模', '建议单笔亏损上限', '止损跳数(例)', '每跳价值', '最大开仓手数', '说明']
ws3.append([])
ws3.append(headers3)
row = ws3.max_row
for col_num, header in enumerate(headers3, 1):
    cell = ws3.cell(row=row, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

position_data = [
    ['5万', '500元', '10跳', '10元/跳', '5手', '新手起步，严格小仓位'],
    ['10万', '1000元', '10跳', '10元/跳', '10手', '稳定后可适当增加'],
    ['20万', '2000元', '10跳', '10元/跳', '20手', '不要一次满仓'],
    ['50万', '5000元', '10跳', '10元/跳', '50手', '可分批建仓'],
    ['100万', '10000元', '10跳', '10元/跳', '100手', '大资金分批次'],
    ['200万', '20000元', '10跳', '10元/跳', '200手', '不要从100手跳到400手'],
    ['500万', '50000元', '10跳', '10元/跳', '500手', '顶级仓位仍需风控'],
]

for row_data in position_data:
    ws3.append(row_data)

# 设置样式
for row in ws3.iter_rows(min_row=11, max_row=ws3.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border
        cell.alignment = center_align

# 添加进阶说明
ws3.append([])
ws3.append(['进阶仓位管理：盈利后如何加仓'])
ws3[ws3.max_row][0].font = Font(bold=True, size=12)
ws3.append(['阶段', '账户规模', '单笔仓位', '加仓节奏'])
row = ws3.max_row
for col_num in range(1, 5):
    cell = ws3.cell(row=row, column=col_num)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.border = thin_border
    cell.alignment = center_align

advance_data = [
    ['起步阶段', '10万 -> 20万', '10手', '每盈利10%加5~10手'],
    ['成长阶段', '20万 -> 50万', '20手', '每盈利10%加10~15手'],
    ['成熟阶段', '50万 -> 100万', '50手', '每盈利10%加20~30手'],
    ['大资金阶段', '100万 -> 500万', '100手', '每盈利10%加30~50手'],
]
for row_data in advance_data:
    ws3.append(row_data)

for row in ws3.iter_rows(min_row=ws3.max_row-3, max_row=ws3.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        cell.alignment = center_align

column_widths3 = [12, 18, 15, 15, 15, 30]
for i, width in enumerate(column_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

# 保存文件
file_path = '/home/ubuntu/low-low-up/Levi大魔王短线动量交易工具包.xlsx'
wb.save(file_path)
print(f'Excel文件已生成: {file_path}')
